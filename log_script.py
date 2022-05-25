from time import time
from openpyxl import Workbook, load_workbook
from datetime import date, datetime, timedelta


def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days), -1, -1):
        yield start_date + timedelta(n)


wb = Workbook()
ws = wb.active
ws.title = "Data"

ws.append(["ID", "Log Description/Name", "Review Data", "Person", "Remarks"])
sh = wb.active

if date.today().weekday() in range(0, 5):
    now = date.today()

log_files = ["/usr/local/liferay/tomcat-7.0.42/logs/catalina.out",
             "/usr/local/liferay/tomcat-7.0.42/logs/localhost_access_log", "/usr/local/liferay/logs/liferay." + str(now) + ".log"]
id = 0
iteration = -1
days_before = 5
for row in sh.iter_rows(min_row=1, min_col=1, max_row=1, max_col=5):
    for cell in row:
        for x in log_files:

            id += 1
            iteration += 1

            if (iteration % 3 == 0):
                days_before -= 1
                ws.append([id, x, now - timedelta(days_before),
                          "Dominik Warzocha", "OK"])
                log_files[2] = "/usr/local/liferay/logs/liferay." + \
                    str(now - timedelta(days_before)) + ".log"
                continue
            else:
                ws.append([id, x, now - timedelta(days_before),
                          "Dominik Warzocha", "OK"])

wb.save(r"\\nas.siseth.com\Projects\Pepsico - PM\Documentation\Audyt\Log Weekly " + str(now.isocalendar().week) + "_" +
        str(now.isocalendar().year) + ".xlsx")
